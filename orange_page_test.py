import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
import sys
import os

# parent directory is in the system path for module imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from orange_page_login import OrangeLoginPage


@pytest.fixture
def browser():
    chrome_service = ChromeService(r"E:\Automation testing\chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    driver.implicitly_wait(10)
    driver.quit()

def read_test_data_from_excel(file_path=r"E:\Automation testing\testdata.xlsx", sheet_name='Sheet1'):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Ensure both username and password are present
            test_data.append((row[1], row[2]))
        else:
            print(f"Skipping invalid row: {row}")

    return test_data


def orange_page_test(browser, username, password):
    login_page = OrangeLoginPage(browser)

    try:
        # Navigate to the Amazon login page
        login_page.navigate_to_login_page()

        # Wait for username input field to be visible
        WebDriverWait(browser, 20).until(EC.visibility_of_element_located(login_page.username_locator))

        login_page.enter_username(username)
        # login_page.continue_username()

        # Wait for password input field to be visible
        WebDriverWait(browser, 20).until(EC.visibility_of_element_located(login_page.password_locator))

        login_page.enter_password(password)
        login_page.click_login_button()

        # Wait for login success
        WebDriverWait(browser, 20).until(EC.title_contains("OrangeHRM"))
    except Exception as e:
        print(f"Error during login process: {e}")
        raise

@pytest.mark.parametrize("username,password", read_test_data_from_excel())
def orange_page_test_data_driven(browser, username, password):
    orange_page_test(browser, username, password)


if __name__ == "__main__":
    pytest.main(["-v", "orange_page_test.py"])